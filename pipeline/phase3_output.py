"""Phase 3 — Normalise, dedup, and export.

Implements:
- E1: Postcodes.io bulk lookup (100 per request)
- E2: Phone normalisation to E.164 via phonenumbers
- E3: Final CID dedup confirmation
- E4: XLSX export via openpyxl
- E5: Phase 3 runner (checkpoint/resume)
"""
from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

import httpx

from database import (
    check_duplicate_cids,
    get_output_status,
    list_output_pending_leads,
    mark_leads_output_done,
    mark_leads_output_failed,
    reset_running_output_leads,
    retry_failed_output_leads,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# E1 — Postcodes.io bulk lookup
# ---------------------------------------------------------------------------

def bulk_postcode_lookup(
    postcodes: list[str],
    *,
    client: httpx.Client | None = None,
    timeout: float = 10.0,
) -> dict[str, dict[str, Any] | None]:
    """Look up up to 100 postcodes per request using Postcodes.io bulk endpoint.

    Returns a dict keyed by postcode → result dict (or None if not found / invalid).
    """
    if not postcodes:
        return {}

    own_client = client is None
    active_client = client or httpx.Client(timeout=timeout)
    results: dict[str, dict[str, Any] | None] = {}

    try:
        for batch_start in range(0, len(postcodes), 100):
            batch = postcodes[batch_start:batch_start + 100]
            try:
                response = active_client.post(
                    "https://api.postcodes.io/postcodes",
                    json={"postcodes": batch},
                )
                response.raise_for_status()
                for item in response.json().get("result") or []:
                    query = item.get("query") or ""
                    results[query] = item.get("result")
            except Exception as exc:
                logger.warning("Postcodes.io batch failed: %s", exc)
                for postcode in batch:
                    results.setdefault(postcode, None)
    finally:
        if own_client:
            active_client.close()

    return results


def _extract_postcode_from_address(address: str | None) -> str | None:
    """Extract a UK postcode from a full address string."""
    if not address:
        return None
    import re
    match = re.search(r"\b([A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2})\b", address, re.IGNORECASE)
    if not match:
        return None
    return re.sub(r"\s+", "", match.group(1)).upper()


# ---------------------------------------------------------------------------
# E2 — Phone normalisation
# ---------------------------------------------------------------------------

def normalise_uk_phone(raw: str | None) -> str | None:
    """Normalise a UK phone number to E.164 format using phonenumbers.

    Returns the E.164 string on success, or the original value on failure.
    Returns None if raw is None.
    """
    if raw is None:
        return None
    try:
        import phonenumbers
        parsed = phonenumbers.parse(raw, "GB")
        if phonenumbers.is_valid_number(parsed):
            return phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
    except Exception:
        pass
    return raw


# ---------------------------------------------------------------------------
# E3 — Dedup confirmation
# ---------------------------------------------------------------------------

def confirm_no_duplicate_cids(project_id: str, db_path: str | Path) -> list[str]:
    """Return any CIDs that appear more than once. Should be empty list in normal operation."""
    duplicates = check_duplicate_cids(project_id, db_path)
    if duplicates:
        logger.warning(
            "Phase 3: %d duplicate CID(s) found for project %s: %s",
            len(duplicates),
            project_id,
            duplicates[:10],
        )
    return duplicates


# ---------------------------------------------------------------------------
# E4 — XLSX export
# ---------------------------------------------------------------------------

EXPORT_HEADERS = [
    "business_name",
    "category",
    "address",
    "postcode",
    "phone_raw",
    "phone_e164",
    "website",
    "email",
    "email_confidence",
    "owner_name",
    "company_number",
    "google_cid",
    "source",
    "scrape_date",
    "lat",
    "lng",
    "ward",
    "local_authority",
]


def _build_export_row(
    lead: dict[str, Any],
    geo_data: dict[str, dict[str, Any] | None],
) -> list[Any]:
    raw_data = json.loads(lead["raw_data"]) if isinstance(lead["raw_data"], str) else (lead.get("raw_data") or {})
    enrichment_raw = lead.get("enrichment_data")
    enrichment = json.loads(enrichment_raw) if isinstance(enrichment_raw, str) and enrichment_raw else {}

    outreach = enrichment.get("outreach") or {}
    company = enrichment.get("company") or {}
    emails = enrichment.get("emails") or []

    phone_raw = raw_data.get("phone") or lead.get("primary_phone")
    phone_e164 = normalise_uk_phone(phone_raw)
    address = raw_data.get("address") or ""
    postcode = _extract_postcode_from_address(address)

    primary_email = outreach.get("primary_email") or lead.get("primary_email")
    email_confidence = next(
        (e.get("confidence") for e in emails if e.get("address") == primary_email),
        None,
    )

    geo = geo_data.get(postcode) if postcode else None
    lat = geo.get("latitude") if geo else raw_data.get("latitude")
    lng = geo.get("longitude") if geo else raw_data.get("longitude")
    ward = geo.get("admin_ward") if geo else None
    local_authority = geo.get("admin_district") if geo else None

    scrape_date = (lead.get("last_updated") or "")[:10] or None

    return [
        raw_data.get("title"),
        raw_data.get("category"),
        address or None,
        postcode,
        phone_raw,
        phone_e164,
        raw_data.get("website"),
        primary_email,
        email_confidence,
        outreach.get("primary_person") or lead.get("primary_person"),
        company.get("companies_house_number"),
        lead.get("cid"),
        lead.get("source", "gosom"),
        scrape_date,
        lat,
        lng,
        ward,
        local_authority,
    ]


def export_leads_xlsx(
    project_id: str,
    output_path: str | Path,
    db_path: str | Path,
    *,
    postcode_lookup: Any | None = None,
) -> Path:
    """Export all output-pending leads to an XLSX file.

    Performs bulk Postcodes.io lookup before the write loop.
    Returns the path to the written file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for export: pip install openpyxl") from exc

    leads = list_output_pending_leads(project_id, db_path)

    # Collect unique postcodes for bulk lookup
    postcodes: list[str] = []
    seen_postcodes: set[str] = set()
    for lead in leads:
        raw = json.loads(lead["raw_data"]) if isinstance(lead["raw_data"], str) else (lead.get("raw_data") or {})
        pc = _extract_postcode_from_address(raw.get("address") or "")
        if pc and pc not in seen_postcodes:
            postcodes.append(pc)
            seen_postcodes.add(pc)

    # E1: Bulk postcode lookup
    if postcode_lookup is not None:
        geo_data = postcode_lookup(postcodes)
    else:
        geo_data = bulk_postcode_lookup(postcodes) if postcodes else {}

    # E3: Confirm no duplicate CIDs
    confirm_no_duplicate_cids(project_id, db_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center")
    for col, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    exported_ids: list[str] = []
    for row_idx, lead in enumerate(leads, 2):
        row_values = _build_export_row(lead, geo_data)
        for col_idx, value in enumerate(row_values, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        exported_ids.append(lead["id"])

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Mark all exported leads as done
    mark_leads_output_done(exported_ids, db_path)

    return output_path


# ---------------------------------------------------------------------------
# E5 — Phase 3 runner
# ---------------------------------------------------------------------------

def run_phase3(
    project_id: str,
    *,
    db_path: str | Path,
    export_dir: str | Path | None = None,
    postcode_lookup: Any | None = None,
) -> dict[str, Any]:
    """Run Phase 3: normalise, dedup, and export for a project.

    Resets stranded running leads, exports all pending leads to xlsx, and
    returns a summary with the export path and counts.
    """
    reset_running_output_leads(project_id, db_path)

    db_path = Path(db_path)
    resolved_export_dir = Path(export_dir) if export_dir else db_path.parent / "exports"
    export_path = resolved_export_dir / f"{project_id}_leads.xlsx"

    try:
        written_path = export_leads_xlsx(
            project_id,
            export_path,
            db_path,
            postcode_lookup=postcode_lookup,
        )
        status = get_output_status(project_id, db_path)
        return {
            "project_id": project_id,
            "phase": "3",
            "status": "done",
            "export_path": str(written_path),
            "output_status": status,
        }
    except Exception as exc:
        logger.error("Phase 3 export failed for project %s: %s", project_id, exc)
        return {
            "project_id": project_id,
            "phase": "3",
            "status": "failed",
            "error": str(exc),
            "output_status": get_output_status(project_id, db_path),
        }
