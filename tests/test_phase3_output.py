"""E1–E6 — Phase 3 output tests.

E1: bulk_postcode_lookup batches ≤100 per request
E2: normalise_uk_phone → E.164
E3: confirm_no_duplicate_cids
E4: export_leads_xlsx writes correct headers and row values
E5: run_phase3 runner + API endpoints
E6: phone formatting, export row count, no duplicate CIDs, file fields
"""
from __future__ import annotations

import json
import uuid
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock, patch

import pytest

from database import (
    check_duplicate_cids,
    create_project,
    get_connection,
    get_output_status,
    insert_leads_from_gosom_results,
    list_output_pending_leads,
    mark_leads_output_done,
    reset_running_output_leads,
    retry_failed_output_leads,
    update_lead_website_enrichment,
)
from pipeline.phase2_enrichment import build_enrichment_record
from pipeline.phase3_output import (
    EXPORT_HEADERS,
    bulk_postcode_lookup,
    confirm_no_duplicate_cids,
    export_leads_xlsx,
    normalise_uk_phone,
    run_phase3,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _raw_lead(cid: str | None = None, postcode_in_address: str = "SA1 1DP") -> dict[str, Any]:
    return {
        "cid": cid or str(uuid.uuid4()),
        "title": "Test Plumbers Ltd",
        "category": "Plumber",
        "address": f"1 High Street, Swansea, {postcode_in_address}",
        "phone": "01792 123456",
        "website": "https://example.co.uk",
        "latitude": 51.6193,
        "longitude": -3.9437,
    }


def _insert_lead_with_enrichment(
    project_id: str,
    db_path: Path,
    raw: dict[str, Any],
    primary_email: str | None = "test@example.co.uk",
) -> str:
    insert_leads_from_gosom_results(
        project_id=project_id, cell_id="cell-1", results=[raw], db_path=db_path
    )
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT id FROM leads WHERE project_id = ? AND cid = ?",
            (project_id, raw["cid"]),
        ).fetchone()
    lead_id = row["id"]
    record = build_enrichment_record(lead_id=lead_id, raw_data=raw)
    enrichment = record.model_dump(mode="json")
    if primary_email:
        enrichment["outreach"]["primary_email"] = primary_email
        enrichment["outreach"]["ready"] = True
    update_lead_website_enrichment(
        lead_id=lead_id, enrichment_data=enrichment, db_path=db_path,
        website_status="done", ai_fallback_status="done", whois_mx_status="done",
    )
    return lead_id


# ---------------------------------------------------------------------------
# E1 — bulk_postcode_lookup
# ---------------------------------------------------------------------------

def test_bulk_postcode_lookup_batches_100_per_request() -> None:
    """More than 100 postcodes should be split into multiple requests."""
    call_count = 0
    received_batches: list[list[str]] = []

    def fake_post_handler(url: str, json: dict, **kwargs) -> MagicMock:
        nonlocal call_count
        call_count += 1
        received_batches.append(json.get("postcodes", []))
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = {"result": [
            {"query": pc, "result": {"latitude": 51.6, "longitude": -3.9, "admin_ward": "Castle", "admin_district": "Swansea"}}
            for pc in json.get("postcodes", [])
        ]}
        return mock_resp

    mock_client = MagicMock()
    mock_client.post.side_effect = fake_post_handler

    postcodes = [f"SA{i}1DP" for i in range(105)]
    bulk_postcode_lookup(postcodes, client=mock_client)

    assert call_count == 2
    assert len(received_batches[0]) == 100
    assert len(received_batches[1]) == 5


def test_bulk_postcode_lookup_returns_empty_for_empty_input() -> None:
    result = bulk_postcode_lookup([])
    assert result == {}


def test_bulk_postcode_lookup_tolerates_api_failure() -> None:
    mock_client = MagicMock()
    mock_client.post.side_effect = Exception("network error")
    result = bulk_postcode_lookup(["SA11DP"], client=mock_client)
    assert "SA11DP" in result
    assert result["SA11DP"] is None


# ---------------------------------------------------------------------------
# E2 — normalise_uk_phone
# ---------------------------------------------------------------------------

def test_normalise_uk_phone_landline() -> None:
    assert normalise_uk_phone("01792 123456") == "+441792123456"


def test_normalise_uk_phone_already_e164() -> None:
    assert normalise_uk_phone("+441792123456") == "+441792123456"


def test_normalise_uk_phone_mobile() -> None:
    result = normalise_uk_phone("07911 123456")
    assert result == "+447911123456"


def test_normalise_uk_phone_no_spaces() -> None:
    assert normalise_uk_phone("01792123456") == "+441792123456"


def test_normalise_uk_phone_garbage_returns_original() -> None:
    assert normalise_uk_phone("garbage") == "garbage"


def test_normalise_uk_phone_none_returns_none() -> None:
    assert normalise_uk_phone(None) is None


# ---------------------------------------------------------------------------
# E3 — confirm_no_duplicate_cids
# ---------------------------------------------------------------------------

def test_confirm_no_duplicate_cids_returns_empty_for_clean_project(temp_db: Path) -> None:
    project = create_project(
        name="Dedup Test", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )
    raw = _raw_lead()
    insert_leads_from_gosom_results(
        project_id=project["id"], cell_id="cell-1", results=[raw], db_path=temp_db
    )
    dups = confirm_no_duplicate_cids(project["id"], temp_db)
    assert dups == []


def test_schema_prevents_duplicate_cids_in_same_project(temp_db: Path) -> None:
    """The UNIQUE(project_id, cid) constraint must block duplicate inserts."""
    project = create_project(
        name="Dup CID", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )
    shared_cid = str(uuid.uuid4())
    raw1 = {"cid": shared_cid, "title": "First"}
    raw2 = {"cid": shared_cid, "title": "Duplicate"}

    result1 = insert_leads_from_gosom_results(
        project_id=project["id"], cell_id="cell-1", results=[raw1], db_path=temp_db
    )
    result2 = insert_leads_from_gosom_results(
        project_id=project["id"], cell_id="cell-1", results=[raw2], db_path=temp_db
    )
    # First insert succeeds; second is a duplicate
    assert result1["inserted"] == 1
    assert result2["inserted"] == 0
    assert result2["duplicates"] == 1

    # check_duplicate_cids confirms the project is clean
    dups = check_duplicate_cids(project["id"], temp_db)
    assert dups == []


# ---------------------------------------------------------------------------
# E4 — export_leads_xlsx
# ---------------------------------------------------------------------------

def test_export_leads_xlsx_produces_file_with_correct_headers(tmp_path: Path, temp_db: Path) -> None:
    import openpyxl

    project = create_project(
        name="Export Test", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )
    raw = _raw_lead()
    _insert_lead_with_enrichment(project["id"], temp_db, raw)

    export_path = tmp_path / "leads.xlsx"
    no_op_postcode_lookup = lambda pcs: {}
    result_path = export_leads_xlsx(
        project["id"], export_path, temp_db, postcode_lookup=no_op_postcode_lookup
    )

    assert result_path.exists()
    wb = openpyxl.load_workbook(result_path)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(EXPORT_HEADERS) + 1)]
    assert headers == EXPORT_HEADERS


def test_export_leads_xlsx_writes_business_name_and_phone(tmp_path: Path, temp_db: Path) -> None:
    import openpyxl

    project = create_project(
        name="Phone Export", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )
    raw = _raw_lead()
    _insert_lead_with_enrichment(project["id"], temp_db, raw)

    export_path = tmp_path / "leads.xlsx"
    export_leads_xlsx(project["id"], export_path, temp_db, postcode_lookup=lambda _: {})

    wb = openpyxl.load_workbook(export_path)
    ws = wb.active
    # Row 2 = first data row
    row_values = {EXPORT_HEADERS[c - 1]: ws.cell(row=2, column=c).value for c in range(1, len(EXPORT_HEADERS) + 1)}
    assert row_values["business_name"] == "Test Plumbers Ltd"
    assert row_values["phone_raw"] == "01792 123456"
    # E.164 normalisation applied
    assert row_values["phone_e164"] == "+441792123456"
    assert row_values["google_cid"] == raw["cid"]


def test_export_leads_xlsx_marks_leads_output_done(tmp_path: Path, temp_db: Path) -> None:
    project = create_project(
        name="Status Export", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )
    raw = _raw_lead()
    _insert_lead_with_enrichment(project["id"], temp_db, raw)

    export_leads_xlsx(
        project["id"], tmp_path / "leads.xlsx", temp_db, postcode_lookup=lambda _: {}
    )

    status = get_output_status(project["id"], temp_db)
    assert status["counts"].get("done", 0) == 1
    assert status["counts"].get("pending", 0) == 0


def test_export_leads_xlsx_row_count_matches_leads(tmp_path: Path, temp_db: Path) -> None:
    import openpyxl

    project = create_project(
        name="Row Count", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )
    raws = [_raw_lead(cid=f"cid-{i}") for i in range(3)]
    for raw in raws:
        _insert_lead_with_enrichment(project["id"], temp_db, raw)

    export_path = tmp_path / "leads.xlsx"
    export_leads_xlsx(project["id"], export_path, temp_db, postcode_lookup=lambda _: {})

    wb = openpyxl.load_workbook(export_path)
    ws = wb.active
    # Row 1 = headers, rows 2..N = data
    data_rows = ws.max_row - 1
    assert data_rows == 3


def test_export_includes_email_and_postcode(tmp_path: Path, temp_db: Path) -> None:
    import openpyxl

    project = create_project(
        name="Email Export", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )
    raw = _raw_lead(postcode_in_address="SA1 1DP")
    _insert_lead_with_enrichment(project["id"], temp_db, raw, primary_email="info@example.co.uk")

    export_path = tmp_path / "leads.xlsx"
    export_leads_xlsx(project["id"], export_path, temp_db, postcode_lookup=lambda _: {})

    wb = openpyxl.load_workbook(export_path)
    ws = wb.active
    row = {EXPORT_HEADERS[c - 1]: ws.cell(row=2, column=c).value for c in range(1, len(EXPORT_HEADERS) + 1)}
    assert row["email"] == "info@example.co.uk"
    assert row["postcode"] == "SA11DP"


# ---------------------------------------------------------------------------
# E5 — run_phase3 + reset/retry lifecycle
# ---------------------------------------------------------------------------

def test_run_phase3_returns_done_status_and_writes_file(tmp_path: Path, temp_db: Path) -> None:
    project = create_project(
        name="Phase3 Run", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )
    raw = _raw_lead()
    _insert_lead_with_enrichment(project["id"], temp_db, raw)

    result = run_phase3(
        project["id"],
        db_path=temp_db,
        export_dir=tmp_path,
        postcode_lookup=lambda _: {},
    )
    assert result["status"] == "done"
    assert Path(result["export_path"]).exists()


def test_reset_running_output_leads_returns_stranded_to_pending(temp_db: Path) -> None:
    project = create_project(
        name="Reset Output", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )
    raw = _raw_lead()
    lead_id = _insert_lead_with_enrichment(project["id"], temp_db, raw)
    with get_connection(temp_db) as conn:
        conn.execute("UPDATE leads SET output_status = 'running' WHERE id = ?", (lead_id,))
        conn.commit()

    reset_count = reset_running_output_leads(project["id"], temp_db)
    assert reset_count == 1

    with get_connection(temp_db) as conn:
        row = conn.execute("SELECT output_status FROM leads WHERE id = ?", (lead_id,)).fetchone()
    assert row["output_status"] == "pending"


def test_retry_failed_output_leads_requeues_failed(temp_db: Path) -> None:
    project = create_project(
        name="Retry Output", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )
    raw = _raw_lead()
    lead_id = _insert_lead_with_enrichment(project["id"], temp_db, raw)
    with get_connection(temp_db) as conn:
        conn.execute("UPDATE leads SET output_status = 'failed' WHERE id = ?", (lead_id,))
        conn.commit()

    retried = retry_failed_output_leads(project["id"], temp_db)
    assert retried == 1

    with get_connection(temp_db) as conn:
        row = conn.execute("SELECT output_status FROM leads WHERE id = ?", (lead_id,)).fetchone()
    assert row["output_status"] == "retry"


# ---------------------------------------------------------------------------
# E5 — API endpoints
# ---------------------------------------------------------------------------

def _build_api_client(db_path: Path):
    from fastapi.testclient import TestClient
    from config import AppSettings
    import main as app_module
    settings = AppSettings(scraper_db_path=db_path)
    return TestClient(app_module.create_app(settings_override=settings))


def test_api_phase3_run_returns_202(temp_db: Path) -> None:
    project_id = create_project(
        name="API Phase3", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )["id"]
    client = _build_api_client(temp_db)
    response = client.post(f"/api/projects/{project_id}/phases/3/run")
    assert response.status_code == 202
    assert response.json()["status"] == "started"


def test_api_phase3_status_returns_output_counts(temp_db: Path) -> None:
    project_id = create_project(
        name="API Phase3 Status", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )["id"]
    client = _build_api_client(temp_db)
    response = client.get(f"/api/projects/{project_id}/phases/3/status")
    assert response.status_code == 200
    assert "output_status" in response.json()


def test_api_leads_export_returns_404_when_no_file(temp_db: Path) -> None:
    project_id = create_project(
        name="No Export", primary_term="plumbers",
        bbox="51.5,51.6,-3.9,-3.8", db_path=temp_db,
    )["id"]
    client = _build_api_client(temp_db)
    response = client.get(f"/api/projects/{project_id}/leads/export")
    assert response.status_code == 404


def test_api_phase3_endpoints_return_404_for_missing_project(temp_db: Path) -> None:
    client = _build_api_client(temp_db)
    missing = "00000000-0000-0000-0000-000000000000"
    assert client.post(f"/api/projects/{missing}/phases/3/run").status_code == 404
    assert client.post(f"/api/projects/{missing}/phases/3/resume").status_code == 404
    assert client.get(f"/api/projects/{missing}/phases/3/status").status_code == 404
    assert client.get(f"/api/projects/{missing}/leads/export").status_code == 404
